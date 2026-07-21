#!/usr/bin/env python3
"""
Vérification ponctuelle — Format des Signal code (Maintenance préventive / Réparation après panne)
=====================================================================================================

Contrôle indépendant de la Validité "Signal reference" déjà faite dans
verify_maintenance_preventive.py (qui ne porte que sur Appel maintenance
préventive) : ici, on vérifie le champ "Signal code" tel que saisi
directement dans les formulaires Maintenance préventive et Réparation
après panne, pour repérer les codes qui ne respectent pas le format
{DEPLOYMENT}_{JJMMAAAA}_{E|S}{N} (ex. MAR_21072026_E14).

Même principe de log que verify_maintenance_preventive.py, mais dans un
fichier séparé (data_verification_signal_code_log.xlsx) : seules les
anomalies sont journalisées, avec un suivi Nouveau / Toujours ouvert /
Résolu à travers les exécutions successives (action déclenchée
manuellement, pas de cron).

Pour les erreurs de formatage simples (espace, tiret/slash au lieu
d'underscore, underscore manquant avant E/S, E/S et numéro inversés,
caractère parasite en fin de code), une valeur corrigée est proposée
automatiquement dans la colonne "Signal code correct". Si l'anomalie
est ambiguë (chiffres manquants/en trop dans la date, code trop
dégradé), la colonne reste vide : correction manuelle nécessaire.
"""

import csv
import io
import os
import re
import sys
from collections import Counter
from datetime import datetime

import requests

MWATER_API_BASE = "https://api.mwater.co/v3"

DATAGRID_MAINTENANCE = "2cfe0ba7ac264d119cfc8964b5f3cebc"
DATAGRID_REPARATION = "d9f1c36a2d6340429658b6628fe81b88"

LOG_FILE_NAME = "data_verification_signal_code_log.xlsx"

LOG_HEADERS = [
    "Formulaire", "Response Code", "Signal code", "Anomalie",
    "Signal code correct", "Statut", "Première détection",
    "Dernière détection", "Date de résolution",
]

SIGNAL_CODE_RE = re.compile(r"^([A-Z]+)_(\d{2})(\d{2})(\d{4})_([ES])(\d+)$")

# Format "propre" mais avec séparateurs/ordre incorrects : ex. "MAR-08052026-E1",
# "MAR _09072026_E2", "MAR_26/02/2026_S14", "FTU_28042026E1"
LOOSE_RE = re.compile(
    r"^\s*([A-Za-z]+)[\s_\-]*(\d{2})[\s/\-]?(\d{2})[\s/\-]?(\d{4})"
    r"[\s_\-]*([EeSs])[\s_\-]*(\d+)\.?\s*$"
)

# Même chose mais avec E/S et numéro inversés : ex. "MAR_11022025_1E"
SWAPPED_RE = re.compile(
    r"^\s*([A-Za-z]+)[\s_\-]*(\d{2})[\s/\-]?(\d{2})[\s/\-]?(\d{4})"
    r"[\s_\-]*(\d+)[\s_\-]*([EeSs])\.?\s*$"
)


def raise_for_status_verbose(response):
    if not response.ok:
        print(f"HTTP {response.status_code} sur {response.url}", file=sys.stderr)
        print(response.text[:2000], file=sys.stderr)
        response.raise_for_status()


# ---------------------------------------------------------------------------
# mWater
# ---------------------------------------------------------------------------

def mwater_login(username, password):
    resp = requests.post(
        f"{MWATER_API_BASE}/clients",
        json={"username": username, "password": password},
        timeout=30,
    )
    raise_for_status_verbose(resp)
    client_id = resp.json().get("client")
    if not client_id:
        raise RuntimeError("Authentification mWater : champ 'client' absent de la réponse")
    return client_id


def download_datagrid(datagrid_id, client_id):
    resp = requests.get(
        f"{MWATER_API_BASE}/datagrids/{datagrid_id}/download",
        params={"client": client_id, "share": "", "extraFilters": "[]", "format": "csv"},
        timeout=120,
    )
    raise_for_status_verbose(resp)
    text = resp.content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def parse_dt(value):
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Correction proposée
# ---------------------------------------------------------------------------

def propose_correction(raw):
    """Tente de reconstruire un Signal code valide à partir d'une valeur mal formatée.
    Retourne (valeur_corrigee_ou_None, note). None si le code est trop ambigu/dégradé
    pour une correction automatique fiable (nécessite une vérification manuelle)."""

    m = LOOSE_RE.match(raw)
    if not m:
        m = SWAPPED_RE.match(raw)
        if m:
            prefix, dd, mm, yyyy, num, es = m.groups()
        else:
            return None, "Format non reconnu automatiquement"
    else:
        prefix, dd, mm, yyyy, es, num = m.groups()

    # Sanité de la date avant de proposer une correction
    try:
        datetime(int(yyyy), int(mm), int(dd))
    except ValueError:
        return None, "Date invalide dans le code (jour/mois incohérents)"

    corrected = f"{prefix.upper()}_{dd}{mm}{yyyy}_{es.upper()}{num}"
    if corrected == raw:
        return None, "Format non reconnu automatiquement"
    return corrected, ""


def check_format(rows, form_label):
    """Retourne la liste des anomalies (dict) pour un formulaire donné."""
    anomalies = []
    for r in rows:
        code = (r.get("Signal code") or "").strip()
        if not code or SIGNAL_CODE_RE.match(code):
            continue  # vide ou déjà conforme : pas une anomalie

        corrected, note = propose_correction(code)
        description = "Signal code ne respecte pas le format attendu"
        if note:
            description += f" ({note})"

        anomalies.append({
            "Formulaire": form_label,
            "Response Code": r.get("Response Code", ""),
            "Signal code": code,
            "Anomalie": description,
            "Signal code correct": corrected or "",
            "Status brut mWater": r.get("Status", ""),
            "Submitted On": r.get("Submitted On", ""),
        })
    return anomalies


# ---------------------------------------------------------------------------
# Log persistant — Nouveau / Toujours ouvert / Résolu (même principe que
# verify_maintenance_preventive.py)
# ---------------------------------------------------------------------------

def anomaly_key(a):
    return (a["Formulaire"], a["Response Code"])


def download_existing_log(token, drive_id, folder_item_id, file_name):
    resp = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{folder_item_id}:/{file_name}:/content",
        headers={"Authorization": f"Bearer {token}"},
        timeout=60,
    )
    if resp.status_code == 404:
        return []
    raise_for_status_verbose(resp)

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["Anomalies"]
    headers = [c.value for c in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if row.get("Formulaire"):
            rows.append(row)
    return rows


def merge_with_log(current_anomalies, existing_rows, today_str):
    existing_open = {}
    already_resolved = []
    for row in existing_rows:
        key = (row.get("Formulaire"), row.get("Response Code"))
        if row.get("Statut") == "Résolu":
            already_resolved.append(row)
        else:
            existing_open[key] = row

    merged = []
    seen_keys = set()
    new_count = 0

    for a in current_anomalies:
        key = anomaly_key(a)
        seen_keys.add(key)
        existing = existing_open.get(key)
        if not existing:
            new_count += 1
        merged.append({
            "Formulaire": a["Formulaire"],
            "Response Code": a["Response Code"],
            "Signal code": a["Signal code"],
            "Anomalie": a["Anomalie"],
            "Signal code correct": a["Signal code correct"],
            "Statut": "Toujours ouvert" if existing else "Nouveau",
            "Première détection": existing.get("Première détection") if existing else today_str,
            "Dernière détection": today_str,
            "Date de résolution": "",
        })

    resolved_count = 0
    submitted_on_lookup = {
        a["Response Code"]: parse_dt(a["Submitted On"]).strftime("%d/%m/%Y")
        for a in current_anomalies if parse_dt(a["Submitted On"])
    }
    for key, row in existing_open.items():
        if key not in seen_keys:
            row = dict(row)
            row["Statut"] = "Résolu"
            row["Date de résolution"] = submitted_on_lookup.get(row.get("Response Code"), today_str)
            merged.append(row)
            resolved_count += 1

    merged.extend(already_resolved)
    return merged, new_count, resolved_count


def build_log_report(merged_rows, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Anomalies"
    ws.append(LOG_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in merged_rows:
        ws.append([row.get(h, "") for h in LOG_HEADERS])

    open_rows = [r for r in merged_rows if r.get("Statut") != "Résolu"]
    summary = wb.create_sheet("Résumé")
    summary.append(["Formulaire", "Nombre d'anomalies ouvertes"])
    for cell in summary[1]:
        cell.font = Font(bold=True)
    counts = Counter(r.get("Formulaire") for r in open_rows)
    for label in ["Maintenance préventive", "Réparation après panne"]:
        summary.append([label, counts.get(label, 0)])
    summary.append(["Total ouvert", len(open_rows)])

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Microsoft Graph
# ---------------------------------------------------------------------------

def graph_token(tenant_id, client_id, client_secret):
    resp = requests.post(
        f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token",
        data={
            "grant_type": "client_credentials",
            "client_id": client_id,
            "client_secret": client_secret,
            "scope": "https://graph.microsoft.com/.default",
        },
        timeout=30,
    )
    raise_for_status_verbose(resp)
    return resp.json()["access_token"]


def upload_to_sharepoint(token, drive_id, folder_item_id, file_path, file_name):
    with open(file_path, "rb") as f:
        content = f.read()
    resp = requests.put(
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{folder_item_id}:/{file_name}:/content",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/octet-stream",
        },
        data=content,
        timeout=120,
    )
    raise_for_status_verbose(resp)
    return resp.json()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    mwater_username = os.environ["MWATER_USERNAME"]
    mwater_password = os.environ["MWATER_PASSWORD"]

    azure_tenant_id = os.environ["AZURE_TENANT_ID"]
    azure_client_id = os.environ["AZURE_CLIENT_ID"]
    azure_client_secret = os.environ["AZURE_CLIENT_SECRET"]

    sharepoint_drive_id = os.environ["SHAREPOINT_DRIVE_ID"]
    sharepoint_folder_item_id = os.environ["SHAREPOINT_FOLDER_ITEM_ID"]

    print("Authentification mWater...")
    client_id = mwater_login(mwater_username, mwater_password)

    print("Téléchargement des datagrids...")
    maintenance_rows = download_datagrid(DATAGRID_MAINTENANCE, client_id)
    reparation_rows = download_datagrid(DATAGRID_REPARATION, client_id)
    print(f"  Maintenance préventive : {len(maintenance_rows)} réponses")
    print(f"  Réparation après panne : {len(reparation_rows)} réponses")

    print("Vérification du format des Signal code...")
    anomalies = []
    anomalies += check_format(maintenance_rows, "Maintenance préventive")
    anomalies += check_format(reparation_rows, "Réparation après panne")
    print(f"  {len(anomalies)} anomalies détectées")

    print("Authentification Microsoft Graph...")
    token = graph_token(azure_tenant_id, azure_client_id, azure_client_secret)

    print("Téléchargement du log existant sur SharePoint...")
    existing_rows = download_existing_log(token, sharepoint_drive_id, sharepoint_folder_item_id, LOG_FILE_NAME)
    print(f"  {len(existing_rows)} lignes déjà présentes dans le log")

    today_str = datetime.now().strftime("%d/%m/%Y")
    merged_rows, new_count, resolved_count = merge_with_log(anomalies, existing_rows, today_str)
    open_rows = [r for r in merged_rows if r.get("Statut") != "Résolu"]
    print(f"  {len(open_rows)} anomalies ouvertes ({new_count} nouvelles, {resolved_count} résolues)")

    output_path = f"/tmp/{LOG_FILE_NAME}"
    build_log_report(merged_rows, output_path)
    print(f"Rapport généré : {output_path}")

    print("Dépôt du log mis à jour sur SharePoint...")
    upload_to_sharepoint(token, sharepoint_drive_id, sharepoint_folder_item_id, output_path, LOG_FILE_NAME)

    print("Terminé.")


if __name__ == "__main__":
    main()
