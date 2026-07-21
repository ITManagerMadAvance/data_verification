#!/usr/bin/env python3
"""
Vérification de données — Appel maintenance préventive
=========================================================

Applique les six dimensions du "Manuel de vérification de données" MadAvance
(Complétude, Promptitude, Validité, Unicité, Cohérence, Fiabilité) à
l'activité "Appel maintenance préventive", en croisant les données de
trois formulaires mWater :

  - Appel maintenance préventive
  - Maintenance préventive
  - Réparation après panne

... et d'un quatrième, utilisé uniquement pour la dimension Fiabilité :

  - Première réhabilitation (au sein du datagrid "Première réhabilitation",
    qui contient aussi d'autres types de travaux)

Le script télécharge les données via l'API mWater (datagrids déjà
configurés dans le portail), applique les règles, puis génère un rapport
Excel horodaté (data_verification_call_JJMMAAAA.xlsx) qu'il dépose sur
SharePoint via Microsoft Graph, et envoie un email de confirmation.

Inspiré du repo `ITManagerMadAvance/mWater_backup` pour les mécanismes
d'authentification mWater et Microsoft Graph.
"""

import csv
import io
import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime

import requests

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

MWATER_API_BASE = "https://api.mwater.co/v3"

# IDs des datagrids mWater (préconfigurés dans le portail, pas des secrets)
DATAGRID_APPEL = "5f443b8a8c144502a304c8c5c24d4f82"
DATAGRID_MAINTENANCE = "2cfe0ba7ac264d119cfc8964b5f3cebc"
DATAGRID_REPARATION = "d9f1c36a2d6340429658b6628fe81b88"
DATAGRID_REHAB = "0638d32971704164b9ac22d549d4818e"

# ID du formulaire "Appel maintenance préventive" (pour les requêtes de réponses brutes)
FORM_APPEL = "c08b3fe26d0f42c084074701f29eb75e"

# Fichier de log unique, mis à jour à chaque exécution (pas de nom horodaté :
# c'est le même fichier qui s'enrichit / se met à jour dans le temps)
LOG_FILE_NAME = "data_verification_call_log.xlsx"

LOG_HEADERS = [
    "Dimension", "Sous-dimension", "Response Code", "Water Point ID",
    "Description", "Détails", "Signal code", "Maintenance préventive",
    "Réparation après panne", "Statut", "Première détection",
    "Dernière détection", "Date de résolution",
]

SIGNAL_CODE_RE = re.compile(r"^([A-Z]+)_(\d{2})(\d{2})(\d{4})_([ES])(\d+)$")

DEPLOYMENT_PREFIX_MAP = {
    "MAR": ["Maroantsetra"],
    "FTU": ["Fort-Dauphin", "Taolagnaro"],
}


def raise_for_status_verbose(response):
    """Comme dans backup_mwater.py : loggue le corps complet en cas d'erreur HTTP."""
    if not response.ok:
        print(f"HTTP {response.status_code} sur {response.url}", file=sys.stderr)
        print(response.text[:2000], file=sys.stderr)
        response.raise_for_status()


# ---------------------------------------------------------------------------
# mWater : authentification et téléchargement des datagrids
# ---------------------------------------------------------------------------

def mwater_login(username, password):
    resp = requests.post(
        f"{MWATER_API_BASE}/clients",
        json={"username": username, "password": password},
        timeout=30,
    )
    raise_for_status_verbose(resp)
    client_id = resp.json().get("client")
    if not client_id:
        raise RuntimeError("Authentification mWater : champ 'client' absent de la réponse")
    return client_id


def download_datagrid(datagrid_id, client_id):
    """Télécharge un datagrid mWater et le retourne comme liste de dict (une par ligne)."""
    resp = requests.get(
        f"{MWATER_API_BASE}/datagrids/{datagrid_id}/download",
        params={"client": client_id, "share": "", "extraFilters": "[]", "format": "csv"},
        timeout=120,
    )
    raise_for_status_verbose(resp)
    # utf-8-sig pour gérer le BOM renvoyé par mWater
    text = resp.content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def fetch_raw_responses_by_code(client_id, form_id, codes, chunk_size=200):
    """Récupère les réponses brutes mWater (endpoint /responses, avant jointure du datagrid)
    pour une liste de 'Response Code'. Sert à distinguer un champ jamais répondu d'un champ
    répondu dont l'entité liée (ex. Water Point) a été supprimée : dans ce cas, le datagrid
    exporte le champ vide alors que la réponse contient bien une référence à une entité,
    visible dans le tableau `entities` de la réponse brute.
    """
    results = {}
    codes = [c for c in codes if c]
    for i in range(0, len(codes), chunk_size):
        chunk = codes[i:i + chunk_size]
        resp = requests.get(
            f"{MWATER_API_BASE}/responses",
            params={
                "client": client_id,
                "filter": json.dumps({"form": form_id, "code": {"$in": chunk}}),
            },
            timeout=60,
        )
        raise_for_status_verbose(resp)
        for item in resp.json():
            results[item.get("code")] = item
    return results


def has_entity_reference(raw_response, entity_type="water_point"):
    if not raw_response:
        return False
    return any(e.get("entityType") == entity_type for e in raw_response.get("entities", []))


# ---------------------------------------------------------------------------
# Utilitaires de parsing
# ---------------------------------------------------------------------------

def parse_dt(value):
    """Parse un horodatage mWater 'AAAA-MM-JJ HH:MM:SS'. Retourne None si vide/invalide."""
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def parse_signal_code_date(code):
    """Extrait la date embarquée (JJMMAAAA) d'un signal code/reference. Retourne un datetime.date ou None."""
    m = SIGNAL_CODE_RE.match((code or "").strip())
    if not m:
        return None
    deployment, dd, mm, yyyy, _, _ = m.groups()
    try:
        return datetime(int(yyyy), int(mm), int(dd)).date()
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Anomalie : structure commune
# ---------------------------------------------------------------------------

class Anomaly:
    def __init__(self, dimension, subdimension, response_code, water_point_id, description, details=""):
        self.dimension = dimension
        self.subdimension = subdimension
        self.response_code = response_code
        self.water_point_id = water_point_id
        self.description = description
        self.details = details

    def as_row(self):
        return [self.dimension, self.subdimension, self.response_code, self.water_point_id,
                self.description, self.details]


# ---------------------------------------------------------------------------
# Dimension : Complétude
# ---------------------------------------------------------------------------

def check_completude(appel_rows, client_id=None):
    anomalies = []
    ambiguous = []  # rows Final avec WP ID vide dans le datagrid : à désambiguïser via l'API brute

    for r in appel_rows:
        status = r.get("Status", "")
        rc = r.get("Response Code", "")
        wp = r.get("Water Point ID > Unique ID", "")
        dont_know = (r.get("Water Point ID (Don't Know)") or "").strip().lower() == "true"

        if status == "Draft":
            anomalies.append(Anomaly(
                "Complétude", "Brouillon non soumis", rc, wp,
                "Brouillon jamais soumis",
                f"Drafted On: {r.get('Drafted On', '')}",
            ))
        elif status == "Final" and not wp.strip():
            if dont_know:
                # Réponse explicite "je ne sais pas" -> réponse valide, pas une anomalie.
                continue
            ambiguous.append(rc)

    resolved = {}
    if client_id and ambiguous:
        raw = fetch_raw_responses_by_code(client_id, FORM_APPEL, ambiguous)
        resolved = {code: has_entity_reference(raw.get(code)) for code in ambiguous}

    for rc in ambiguous:
        if resolved.get(rc):
            # La réponse contient bien une référence à un Water Point, mais le datagrid
            # l'exporte vide -> le site lié a probablement été supprimé/est inaccessible.
            anomalies.append(Anomaly(
                "Complétude", "Site associé supprimé/inaccessible", rc, "",
                "Water Point ID répondu mais site associé supprimé ou inaccessible",
                "Non résolu dans l'export mWater malgré une réponse valide (à vérifier dans le portail)",
            ))
        else:
            anomalies.append(Anomaly(
                "Complétude", "Water Point ID manquant", rc, "",
                "Réponse finalisée sans Water Point ID renseigné",
            ))
    return anomalies


# ---------------------------------------------------------------------------
# Dimension : Promptitude
# ---------------------------------------------------------------------------

def check_promptitude(appel_rows):
    anomalies = []
    for r in appel_rows:
        drafted = parse_dt(r.get("Drafted On"))
        submitted = parse_dt(r.get("Submitted On"))
        if drafted and submitted and submitted < drafted:
            anomalies.append(Anomaly(
                "Promptitude", "Chronologie Drafted/Submitted",
                r.get("Response Code", ""), r.get("Water Point ID > Unique ID", ""),
                "Submitted On antérieur à Drafted On",
                f"Drafted On: {drafted} / Submitted On: {submitted}",
            ))
    return anomalies


# ---------------------------------------------------------------------------
# Dimension : Validité
# ---------------------------------------------------------------------------

def check_validite(appel_rows):
    anomalies = []
    for r in appel_rows:
        rc = r.get("Response Code", "")
        wp = r.get("Water Point ID > Unique ID", "")
        signal_ref = r.get("Signal reference", "").strip()
        if signal_ref and not SIGNAL_CODE_RE.match(signal_ref):
            anomalies.append(Anomaly(
                "Validité", "Format signal code", rc, wp,
                "Signal reference ne respecte pas le format attendu",
                f"Valeur : '{signal_ref}'",
            ))
        if wp and not wp.strip().isdigit():
            anomalies.append(Anomaly(
                "Validité", "Format Water Point ID", rc, wp,
                "Water Point ID non numérique",
            ))
    return anomalies


# ---------------------------------------------------------------------------
# Dimension : Unicité
# ---------------------------------------------------------------------------

def check_unicite(appel_rows):
    anomalies = []
    counts = Counter(
        r.get("Signal reference", "").strip()
        for r in appel_rows
        if r.get("Signal reference", "").strip()
    )
    duplicated_codes = {code for code, n in counts.items() if n > 1}
    for r in appel_rows:
        signal_ref = r.get("Signal reference", "").strip()
        status = r.get("Status")
        rejection = (r.get("Rejection message") or "").strip()
        if signal_ref in duplicated_codes:
            anomalies.append(Anomaly(
                "Unicité", "Doublon signal code",
                r.get("Response Code", ""), r.get("Water Point ID > Unique ID", ""),
                "Signal reference dupliqué",
                signal_ref,
            ))
        elif status == "Rejected" and "doublon" in rejection.lower():
            # Uniquement si toujours au statut Rejected : si la réponse est devenue Final,
            # le rejet a déjà été corrigé et n'est plus une anomalie à traiter.
            anomalies.append(Anomaly(
                "Unicité", "Doublon signal code",
                r.get("Response Code", ""), r.get("Water Point ID > Unique ID", ""),
                "Rejet mWater pour doublon de signal code",
                signal_ref if signal_ref else "Signal reference vide",
            ))
    return anomalies


# ---------------------------------------------------------------------------
# Dimension : Cohérence
# ---------------------------------------------------------------------------

def index_by_signal_code(rows, field="Signal code"):
    index = defaultdict(list)
    for r in rows:
        code = (r.get(field) or "").strip()
        if code:
            index[code].append(r)
    return index


def check_coherence(appel_rows, maintenance_rows, reparation_rows):
    anomalies = []
    maintenance_idx = index_by_signal_code(maintenance_rows)
    reparation_idx = index_by_signal_code(reparation_rows)

    for r in appel_rows:
        rc = r.get("Response Code", "")
        wp = r.get("Water Point ID > Unique ID", "")
        signal_ref = r.get("Signal reference", "").strip()
        deployment = (r.get("Deployment") or "").strip()
        pump_state = (r.get("Is the pump currently working ?") or "").strip()

        if not signal_ref:
            continue

        m = SIGNAL_CODE_RE.match(signal_ref)
        if not m:
            continue  # déjà signalé en Validité
        code_prefix = m.group(1)

        # 1. Préfixe de déploiement
        if deployment and code_prefix != deployment:
            anomalies.append(Anomaly(
                "Cohérence", "Préfixe déploiement", rc, wp,
                "Préfixe du signal code différent du déploiement déclaré",
                f"Signal reference : {signal_ref} / Deployment : {deployment}",
            ))

        # 2. Présence dans le bon formulaire selon l'état de la pompe
        found_in_maintenance = signal_ref in maintenance_idx
        found_in_reparation = signal_ref in reparation_idx

        if pump_state == "Partially" and not found_in_maintenance:
            anomalies.append(Anomaly(
                "Cohérence", "Présence dans le bon formulaire", rc, wp,
                "Pompe 'Partiellement' fonctionnelle mais signal code absent du formulaire Maintenance préventive",
                f"Signal reference : {signal_ref}",
            ))
        elif pump_state == "No" and not found_in_reparation:
            anomalies.append(Anomaly(
                "Cohérence", "Présence dans le bon formulaire", rc, wp,
                "Pompe 'Non' fonctionnelle mais signal code absent du formulaire Réparation après panne",
                f"Signal reference : {signal_ref}",
            ))

        # 3. Date du signal code <= date de l'activité (Completion date of the work)
        code_date = parse_signal_code_date(signal_ref)
        if code_date is None:
            continue

        matches = maintenance_idx.get(signal_ref, []) + reparation_idx.get(signal_ref, [])
        for match in matches:
            completion = parse_dt(match.get("Completion date of the work"))
            if completion and code_date > completion.date():
                anomalies.append(Anomaly(
                    "Cohérence", "Date signal code vs activité", rc, wp,
                    "Date du signal code postérieure à la date de l'activité",
                    f"Signal code : {signal_ref} (date : {code_date}) / "
                    f"Completion date of the work : {completion.date()}",
                ))
    return anomalies


# ---------------------------------------------------------------------------
# Dimension : Fiabilité
# ---------------------------------------------------------------------------

def check_fiabilite(appel_rows, rehab_rows):
    anomalies = []

    wp_field = "De quel point d'eau s'agit-il? > Unique ID"

    # Sous-ensemble "Première réhabilitation" (le datagrid contient d'autres types de travaux)
    rehab_only = [r for r in rehab_rows if (r.get("Type de travaux") or "").strip() == "Première réhabilitation"]

    # Source de vérité : réhabilitations réussies (Final, pas Rejected)
    success_wp_ids = {
        r.get(wp_field, "").strip()
        for r in rehab_only
        if r.get("Status") == "Final" and r.get(wp_field, "").strip()
    }

    # Contexte complet (tous types de travaux, tous statuts) pour diagnostic
    all_wp_context = defaultdict(list)
    for r in rehab_rows:
        wp = r.get(wp_field, "").strip()
        if wp:
            all_wp_context[wp].append((r.get("Type de travaux", ""), r.get("Status", "")))

    for r in appel_rows:
        rejection = (r.get("Rejection message") or "").strip()
        status = r.get("Status")
        if status != "Rejected" or "id incorrect" not in rejection.lower():
            # Comme pour Unicité : si la réponse est redevenue Final, le rejet a déjà
            # été corrigé et n'est plus une anomalie à traiter.
            continue

        rc = r.get("Response Code", "")
        wp = r.get("Water Point ID > Unique ID", "").strip()

        if wp in success_wp_ids:
            anomalies.append(Anomaly(
                "Fiabilité", "Rejet ID incorrect", rc, wp,
                "Rejet 'ID incorrect' potentiellement injustifié",
                f"Ce Water Point ID existe dans la liste des Premières réhabilitations réussies (Status Final)",
            ))
        elif wp in all_wp_context:
            context = "; ".join(f"{t or 'type inconnu'} ({s})" for t, s in all_wp_context[wp])
            anomalies.append(Anomaly(
                "Fiabilité", "Rejet ID incorrect", rc, wp,
                "Rejet 'ID incorrect' à réexaminer — WP ID trouvé ailleurs dans le datagrid Réhabilitation",
                f"Trouvé sous : {context}",
            ))
        # Sinon : WP ID introuvable dans le datagrid réhabilitation -> rejet probablement justifié, pas d'anomalie.

    return anomalies


# ---------------------------------------------------------------------------
# Rapport Excel — log unique avec suivi Nouveau / Toujours ouvert / Résolu
# ---------------------------------------------------------------------------

def anomaly_key(a):
    """Identifiant stable d'une anomalie à travers les exécutions successives."""
    return (a.dimension, a.subdimension, a.response_code, a.description)


def download_existing_log(token, drive_id, folder_item_id, file_name):
    """Télécharge le log existant sur SharePoint. Retourne [] si le fichier n'existe pas encore
    (première exécution)."""
    resp = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{folder_item_id}:/{file_name}:/content",
        headers={"Authorization": f"Bearer {token}"},
        timeout=60,
    )
    if resp.status_code == 404:
        return []
    raise_for_status_verbose(resp)

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["Anomalies"]
    headers = [c.value for c in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if row.get("Dimension"):  # ignore lignes vides éventuelles
            rows.append(row)
    return rows


def build_submitted_on_lookup(appel_rows):
    """Associe chaque Response Code à son dernier Submitted On (format JJ/MM/AAAA),
    utilisé comme date de résolution réelle plutôt que la date d'exécution du script."""
    lookup = {}
    for r in appel_rows:
        rc = r.get("Response Code")
        dt = parse_dt(r.get("Submitted On"))
        if rc and dt:
            lookup[rc] = dt.strftime("%d/%m/%Y")
    return lookup


def build_signal_code_lookups(appel_rows, maintenance_rows, reparation_rows):
    """Pour chaque Response Code (Appel), donne son Signal reference, et deux index par
    signal code pour savoir si ce code est présent dans Maintenance préventive / Réparation
    après panne — utilisé pour enrichir chaque ligne du log avec ce contexte de croisement."""
    signal_by_rc = {
        r.get("Response Code"): (r.get("Signal reference") or "").strip()
        for r in appel_rows
    }
    maint_idx = index_by_signal_code(maintenance_rows)
    rep_idx = index_by_signal_code(reparation_rows)
    return signal_by_rc, maint_idx, rep_idx


def merge_with_log(current_anomalies, existing_rows, today_str, submitted_on_lookup=None,
                    signal_lookups=None):
    """Fusionne les anomalies détectées aujourd'hui avec le log existant :
    - toujours présente -> Statut 'Toujours ouvert', Dernière détection mise à jour
    - nouvelle -> Statut 'Nouveau'
    - présente avant mais plus détectée aujourd'hui -> Statut 'Résolu', avec pour date de
      résolution le Submitted On de la réponse si connu (date réelle de la correction dans
      mWater), sinon la date d'exécution du script en repli
    - déjà marquée 'Résolu' avant -> conservée telle quelle (historique)
    """
    submitted_on_lookup = submitted_on_lookup or {}
    signal_by_rc, maint_idx, rep_idx = signal_lookups or ({}, {}, {})
    existing_open = {}
    already_resolved = []
    for row in existing_rows:
        key = (row.get("Dimension"), row.get("Sous-dimension"), row.get("Response Code"), row.get("Description"))
        if row.get("Statut") == "Résolu":
            already_resolved.append(row)
        else:
            existing_open[key] = row

    merged = []
    seen_keys = set()
    new_count = 0

    for a in current_anomalies:
        key = anomaly_key(a)
        seen_keys.add(key)
        existing = existing_open.get(key)
        if not existing:
            new_count += 1
        code = signal_by_rc.get(a.response_code, "")
        merged.append({
            "Dimension": a.dimension,
            "Sous-dimension": a.subdimension,
            "Response Code": a.response_code,
            "Water Point ID": a.water_point_id,
            "Description": a.description,
            "Détails": a.details,
            "Signal code": code,
            "Maintenance préventive": "Oui" if code and code in maint_idx else "Non",
            "Réparation après panne": "Oui" if code and code in rep_idx else "Non",
            "Statut": "Toujours ouvert" if existing else "Nouveau",
            "Première détection": existing.get("Première détection") if existing else today_str,
            "Dernière détection": today_str,
            "Date de résolution": "",
        })

    resolved_count = 0
    for key, row in existing_open.items():
        if key not in seen_keys:
            row = dict(row)
            row["Statut"] = "Résolu"
            row["Date de résolution"] = submitted_on_lookup.get(row.get("Response Code"), today_str)
            merged.append(row)
            resolved_count += 1

    merged.extend(already_resolved)
    return merged, new_count, resolved_count


def build_log_report(merged_rows, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Anomalies"
    ws.append(LOG_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in merged_rows:
        ws.append([row.get(h, "") for h in LOG_HEADERS])

    # Onglet résumé : uniquement les anomalies actuellement ouvertes (Nouveau + Toujours ouvert)
    open_rows = [r for r in merged_rows if r.get("Statut") != "Résolu"]
    summary = wb.create_sheet("Résumé")
    summary.append(["Dimension", "Nombre d'anomalies ouvertes"])
    for cell in summary[1]:
        cell.font = Font(bold=True)
    counts = Counter(r.get("Dimension") for r in open_rows)
    for dimension in ["Complétude", "Promptitude", "Validité", "Unicité", "Cohérence", "Fiabilité"]:
        summary.append([dimension, counts.get(dimension, 0)])
    summary.append(["Total ouvert", len(open_rows)])

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Microsoft Graph : upload SharePoint + email
# ---------------------------------------------------------------------------

def graph_token(tenant_id, client_id, client_secret):
    resp = requests.post(
        f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token",
        data={
            "grant_type": "client_credentials",
            "client_id": client_id,
            "client_secret": client_secret,
            "scope": "https://graph.microsoft.com/.default",
        },
        timeout=30,
    )
    raise_for_status_verbose(resp)
    return resp.json()["access_token"]


def upload_to_sharepoint(token, drive_id, folder_item_id, file_path, file_name):
    with open(file_path, "rb") as f:
        content = f.read()
    resp = requests.put(
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{folder_item_id}:/{file_name}:/content",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/octet-stream",
        },
        data=content,
        timeout=120,
    )
    raise_for_status_verbose(resp)
    return resp.json()


def send_confirmation_email(token, sender, recipients, file_name, open_count, new_count,
                             resolved_count, counts_by_dimension):
    lines = "".join(f"<li>{dim} : {n}</li>" for dim, n in counts_by_dimension.items())
    body_html = (
        f"<p>Vérification de données \"Appel maintenance préventive\" exécutée.</p>"
        f"<p>Anomalies actuellement ouvertes : <b>{open_count}</b> "
        f"(dont {new_count} nouvelles) — {resolved_count} résolue(s) depuis la dernière exécution.</p>"
        f"<ul>{lines}</ul>"
        f"<p>Log mis à jour sur SharePoint : <b>{file_name}</b></p>"
    )
    resp = requests.post(
        f"https://graph.microsoft.com/v1.0/users/{sender}/sendMail",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "message": {
                "subject": f"Vérification de données — {file_name}",
                "body": {"contentType": "HTML", "content": body_html},
                "toRecipients": [{"emailAddress": {"address": addr.strip()}} for addr in recipients.split(",")],
            }
        },
        timeout=30,
    )
    raise_for_status_verbose(resp)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    mwater_username = os.environ["MWATER_USERNAME"]
    mwater_password = os.environ["MWATER_PASSWORD"]

    azure_tenant_id = os.environ["AZURE_TENANT_ID"]
    azure_client_id = os.environ["AZURE_CLIENT_ID"]
    azure_client_secret = os.environ["AZURE_CLIENT_SECRET"]

    sharepoint_drive_id = os.environ["SHAREPOINT_DRIVE_ID"]
    sharepoint_folder_item_id = os.environ["SHAREPOINT_FOLDER_ITEM_ID"]

    email_sender = os.environ["EMAIL_SENDER"]
    email_recipients = os.environ["EMAIL_RECIPIENTS"]

    print("Authentification mWater...")
    client_id = mwater_login(mwater_username, mwater_password)

    print("Téléchargement des datagrids...")
    appel_rows = download_datagrid(DATAGRID_APPEL, client_id)
    maintenance_rows = download_datagrid(DATAGRID_MAINTENANCE, client_id)
    reparation_rows = download_datagrid(DATAGRID_REPARATION, client_id)
    rehab_rows = download_datagrid(DATAGRID_REHAB, client_id)
    print(f"  Appel: {len(appel_rows)} / Maintenance: {len(maintenance_rows)} / "
          f"Réparation: {len(reparation_rows)} / Réhabilitation: {len(rehab_rows)}")

    print("Application des règles de vérification...")
    anomalies = []
    anomalies += check_completude(appel_rows, client_id=client_id)
    anomalies += check_promptitude(appel_rows)
    anomalies += check_validite(appel_rows)
    anomalies += check_unicite(appel_rows)
    anomalies += check_coherence(appel_rows, maintenance_rows, reparation_rows)
    anomalies += check_fiabilite(appel_rows, rehab_rows)
    print(f"  {len(anomalies)} anomalies détectées")

    print("Authentification Microsoft Graph...")
    token = graph_token(azure_tenant_id, azure_client_id, azure_client_secret)

    print("Téléchargement du log existant sur SharePoint...")
    existing_rows = download_existing_log(token, sharepoint_drive_id, sharepoint_folder_item_id, LOG_FILE_NAME)
    print(f"  {len(existing_rows)} lignes déjà présentes dans le log")

    today_str = datetime.now().strftime("%d/%m/%Y")
    submitted_on_lookup = build_submitted_on_lookup(appel_rows)
    signal_lookups = build_signal_code_lookups(appel_rows, maintenance_rows, reparation_rows)
    merged_rows, new_count, resolved_count = merge_with_log(
        anomalies, existing_rows, today_str,
        submitted_on_lookup=submitted_on_lookup, signal_lookups=signal_lookups)
    open_rows = [r for r in merged_rows if r.get("Statut") != "Résolu"]
    print(f"  {len(open_rows)} anomalies ouvertes ({new_count} nouvelles, {resolved_count} résolues)")

    output_path = f"/tmp/{LOG_FILE_NAME}"
    build_log_report(merged_rows, output_path)
    print(f"Rapport généré : {output_path}")

    print("Dépôt du log mis à jour sur SharePoint...")
    upload_to_sharepoint(token, sharepoint_drive_id, sharepoint_folder_item_id, output_path, LOG_FILE_NAME)

    print("Envoi de l'email de confirmation...")
    counts_by_dimension = Counter(r.get("Dimension") for r in open_rows)
    send_confirmation_email(token, email_sender, email_recipients, LOG_FILE_NAME,
                             len(open_rows), new_count, resolved_count, counts_by_dimension)

    print("Terminé.")


if __name__ == "__main__":
    main()
